import argparse
import json
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def to_iso_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return None
    return text


def to_float(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def to_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def find_header_row(rows, header_name):
    for idx, row in enumerate(rows):
        if row and str(row[0]).strip() == header_name:
            return idx
    return -1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True)
    parser.add_argument("--replace", action="store_true")
    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    db_path = Path(__file__).resolve().parent.parent / "data.sqlite"
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.executescript(
        """
        CREATE TABLE IF NOT EXISTS projects (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          name TEXT NOT NULL,
          market_date TEXT,
          budget REAL NOT NULL DEFAULT 0,
          notes TEXT,
          created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS sku (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          code TEXT NOT NULL UNIQUE,
          name TEXT NOT NULL,
          source_link TEXT,
          category TEXT,
          sample_status TEXT,
          decision TEXT,
          notes TEXT,
          reorder_point INTEGER NOT NULL DEFAULT 5,
          scale_up_threshold INTEGER NOT NULL DEFAULT 10,
          slow_threshold INTEGER NOT NULL DEFAULT 1,
          is_active INTEGER NOT NULL DEFAULT 1,
          created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS accessory (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          project_id INTEGER NOT NULL,
          description TEXT NOT NULL,
          amount REAL NOT NULL,
          created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS inventory_ledger (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          project_id INTEGER,
          sku_id INTEGER NOT NULL,
          movement_type TEXT NOT NULL CHECK(movement_type IN ('ORDERED', 'ARRIVED', 'SOLD', 'ADJUST')),
          quantity INTEGER NOT NULL,
          unit_cost REAL DEFAULT 0,
          unit_price REAL DEFAULT 0,
          event_date TEXT NOT NULL DEFAULT (date('now')),
          notes TEXT,
          created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        """
    )

    if args.replace:
        cur.execute("DELETE FROM inventory_ledger")
        cur.execute("DELETE FROM accessory")
        cur.execute("DELETE FROM sku")
        cur.execute("DELETE FROM projects")
        conn.commit()

    wb = load_workbook(file_path, data_only=True)

    project_ids = []
    project_sheet = wb["Project"]
    project_rows = list(project_sheet.iter_rows(values_only=True))
    p_start = find_header_row(project_rows, "Project Name")
    if p_start >= 0:
        for row in project_rows[p_start + 1 :]:
            name = row[0] if len(row) > 0 else None
            if not name:
                continue
            market_date = to_iso_date(row[1] if len(row) > 1 else None)
            budget = to_float(row[2] if len(row) > 2 else 0)
            note_cost = to_float(row[3] if len(row) > 3 else 0)
            note_rev = to_float(row[5] if len(row) > 5 else 0)
            notes = f"Imported from Project sheet. Costs per market={note_cost}, Revenue per market={note_rev}"
            cur.execute(
                "INSERT INTO projects(name, market_date, budget, notes) VALUES(?, ?, ?, ?)",
                (str(name).strip(), market_date, budget, notes),
            )
            project_ids.append(cur.lastrowid)

    target_project_id = project_ids[-1] if project_ids else None

    accessory_count = 0
    accessory_sheet = wb["Accessory"]
    accessory_rows = list(accessory_sheet.iter_rows(values_only=True))
    a_start = find_header_row(accessory_rows, "Product Name")
    if a_start >= 0 and target_project_id:
        for row in accessory_rows[a_start + 1 :]:
            product = row[0] if len(row) > 0 else None
            if not product:
                continue
            category = row[2] if len(row) > 2 else None
            cost = to_float(row[3] if len(row) > 3 else 0)
            desc = str(product).strip() if not category else f"{str(product).strip()} ({str(category).strip()})"
            cur.execute(
                "INSERT INTO accessory(project_id, description, amount) VALUES (?, ?, ?)",
                (target_project_id, desc, cost),
            )
            accessory_count += 1

    sku_count = 0
    movement_count = 0
    sku_sheet = wb["SKU"]
    sku_rows = list(sku_sheet.iter_rows(values_only=True))
    s_start = find_header_row(sku_rows, "SKU ID")
    if s_start >= 0:
        for row in sku_rows[s_start + 1 :]:
            sku_code = row[0] if len(row) > 0 else None
            if not sku_code:
                continue

            name = row[1] if len(row) > 1 else ""
            source_link = row[2] if len(row) > 2 else None
            category = row[3] if len(row) > 3 else None
            unit_cost = to_float(row[4] if len(row) > 4 else 0)
            arrival_qty = to_int(row[5] if len(row) > 5 else 0)
            purchased_qty = to_int(row[6] if len(row) > 6 else 0)
            total_price = to_float(row[7] if len(row) > 7 else 0)
            sample_status = row[8] if len(row) > 8 else None
            decision = row[9] if len(row) > 9 else None
            sold_qty = to_int(row[10] if len(row) > 10 else 0)
            notes = row[13] if len(row) > 13 else None
            avg_sold_price = to_float(row[14] if len(row) > 14 else 0)

            scale_threshold = 10
            slow_threshold = 1
            if isinstance(decision, str) and decision.lower() == "scale":
                scale_threshold = max(1, sold_qty)
            if sold_qty == 0:
                slow_threshold = 1

            cur.execute(
                """
                INSERT INTO sku(
                  code, name, source_link, category, sample_status, decision, notes,
                  reorder_point, scale_up_threshold, slow_threshold
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(code) DO UPDATE SET
                  name=excluded.name,
                  source_link=excluded.source_link,
                  category=excluded.category,
                  sample_status=excluded.sample_status,
                  decision=excluded.decision,
                  notes=excluded.notes,
                  scale_up_threshold=excluded.scale_up_threshold,
                  slow_threshold=excluded.slow_threshold
                """,
                (
                    str(sku_code).strip(),
                    str(name).strip(),
                    str(source_link).strip() if source_link else None,
                    str(category).strip() if category else None,
                    str(sample_status).strip() if sample_status else None,
                    str(decision).strip() if decision else None,
                    str(notes).strip() if notes else None,
                    5,
                    scale_threshold,
                    slow_threshold,
                ),
            )

            cur.execute("SELECT id FROM sku WHERE code = ?", (str(sku_code).strip(),))
            sku_id = cur.fetchone()["id"]
            sku_count += 1

            if purchased_qty > 0:
                cur.execute(
                    """
                    INSERT INTO inventory_ledger(project_id, sku_id, movement_type, quantity, unit_cost, unit_price, notes)
                    VALUES (?, ?, 'ORDERED', ?, ?, 0, ?)
                    """,
                    (target_project_id, sku_id, purchased_qty, unit_cost, "Imported from SKU sheet"),
                )
                movement_count += 1
            if arrival_qty > 0:
                cur.execute(
                    """
                    INSERT INTO inventory_ledger(project_id, sku_id, movement_type, quantity, unit_cost, unit_price, notes)
                    VALUES (?, ?, 'ARRIVED', ?, ?, 0, ?)
                    """,
                    (target_project_id, sku_id, arrival_qty, unit_cost, "Imported from SKU sheet"),
                )
                movement_count += 1
            if sold_qty > 0:
                unit_price = avg_sold_price if avg_sold_price > 0 else (total_price / sold_qty if sold_qty else 0)
                cur.execute(
                    """
                    INSERT INTO inventory_ledger(project_id, sku_id, movement_type, quantity, unit_cost, unit_price, notes)
                    VALUES (?, ?, 'SOLD', ?, ?, ?, ?)
                    """,
                    (target_project_id, sku_id, sold_qty, unit_cost, unit_price, "Imported from SKU sheet"),
                )
                movement_count += 1

    conn.commit()
    conn.close()

    print(
        json.dumps(
            {
                "ok": True,
                "file": str(file_path),
                "projectsImported": len(project_ids),
                "accessoriesImported": accessory_count,
                "skuImported": sku_count,
                "movementsImported": movement_count,
                "targetProjectIdForInventory": target_project_id,
            }
        )
    )


if __name__ == "__main__":
    main()
