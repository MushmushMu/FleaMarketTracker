import argparse
import json
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    db_path = Path(__file__).resolve().parent.parent / "data.sqlite"
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)

    project_sheet = wb.create_sheet("Project")
    project_sheet.append(["Project Name", "Market Date", "Budget", "Costs per market", "", "Revenue per market"])
    for cell in project_sheet[1]:
        cell.font = Font(bold=True)

    projects = cur.execute(
        "SELECT name, market_date, budget, notes FROM projects ORDER BY id ASC"
    ).fetchall()
    for project in projects:
        project_sheet.append([project["name"], project["market_date"], project["budget"], "", "", ""])

    accessory_sheet = wb.create_sheet("Accessory")
    accessory_sheet.append(["Product Name", "Source Link", "Category", "Cost"])
    for cell in accessory_sheet[1]:
        cell.font = Font(bold=True)

    accessories = cur.execute(
        """
        SELECT description, category, amount
        FROM accessory
        ORDER BY id ASC
        """
    ).fetchall()
    for row in accessories:
        accessory_sheet.append([row["description"], "", row["category"] or "", row["amount"]])

    sku_sheet = wb.create_sheet("SKU")
    sku_headers = [
        "SKU ID",
        "Product Name",
        "Source Link",
        "Category",
        "Unit Cost",
        "Arrival Quantity",
        "Purchased",
        "Total Price",
        "Sample Evaluation",
        "Decision",
        "Quantity Sold",
        "",
        "",
        "Notes",
        "Average sold price",
        "Current Stock"
    ]
    sku_sheet.append(sku_headers)
    for cell in sku_sheet[1]:
        cell.font = Font(bold=True)

    sku_rows = cur.execute(
        """
        SELECT
          s.code,
          s.name,
          s.source_link,
          s.category,
          s.sample_status,
          s.decision,
          s.notes,
          COALESCE(SUM(CASE WHEN l.movement_type = 'ARRIVED' THEN l.quantity ELSE 0 END), 0) AS arrival_qty,
          COALESCE(SUM(CASE WHEN l.movement_type = 'ORDERED' THEN l.quantity ELSE 0 END), 0) AS purchased_qty,
          COALESCE(SUM(CASE WHEN l.movement_type = 'SOLD' THEN l.quantity ELSE 0 END), 0) AS sold_units,
          COALESCE(SUM(CASE WHEN l.movement_type = 'SOLD' THEN l.quantity * l.unit_price ELSE 0 END), 0) AS revenue,
          COALESCE(
            SUM(CASE WHEN l.movement_type IN ('ORDERED', 'ARRIVED') THEN l.quantity * l.unit_cost ELSE 0 END)
            / NULLIF(SUM(CASE WHEN l.movement_type IN ('ORDERED', 'ARRIVED') THEN l.quantity ELSE 0 END), 0),
            0
          ) AS unit_cost,
          COALESCE(SUM(CASE WHEN l.movement_type = 'ORDERED' THEN l.quantity * l.unit_cost ELSE 0 END), 0) AS total_price,
          COALESCE(SUM(CASE
            WHEN l.movement_type = 'ARRIVED' THEN l.quantity
            WHEN l.movement_type = 'ADJUST' THEN l.quantity
            WHEN l.movement_type = 'SOLD' THEN -l.quantity
            ELSE 0
          END), 0) AS current_stock
        FROM sku s
        LEFT JOIN inventory_ledger l ON l.sku_id = s.id
        GROUP BY s.id
        ORDER BY
          CASE WHEN s.code GLOB 'SKU-[0-9]*' THEN CAST(substr(s.code, 5) AS INTEGER) ELSE 999999 END,
          s.code ASC
        """
    ).fetchall()

    for row in sku_rows:
        sold_units = row["sold_units"] or 0
        revenue = row["revenue"] or 0
        avg_sold_price = revenue / sold_units if sold_units > 0 else 0
        sku_sheet.append([
            row["code"],
            row["name"],
            row["source_link"] or "",
            row["category"] or "",
            row["unit_cost"] or 0,
            row["arrival_qty"] or 0,
            row["purchased_qty"] or 0,
            row["total_price"] or 0,
            row["sample_status"] or "",
            row["decision"] or "",
            sold_units,
            "",
            "",
            row["notes"] or "",
            avg_sold_price,
            row["current_stock"] or 0
        ])

    conn.close()
    wb.save(output_path)

    print(json.dumps({"ok": True, "file": str(output_path)}))


if __name__ == "__main__":
    main()
